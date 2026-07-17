"""
pdf_packager.py
Per-WI PDF packager. OS-agnostic - no Office / COM required.

Builds one merged PDF per work item:
  Page 1+: cover (title, lane, description, AC, comments, inline images)
  Page N+: every attachment converted/included in NAME ASC order

Failures become 1-page error placeholder pages so the final PDF always
has a slot for every input file (strict mode preserved).

Public entry point:
    package_for_wi(wi_dir: Path, paper_size: str = "A4",
                   output_pdf: Path | None = None) -> PackageResult
"""

from __future__ import annotations

import gc
import json
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4, LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Image as RLImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
)

from ado.extract import html_to_text
from tools.office_convert import convert_to_pdf, is_office_extension

# =====================================================================
# Internal contracts
# =====================================================================
PACKET_PDF_NAME: Final[str] = "WI_{wi_id}.pdf"
COVER_PDF_NAME: Final[str] = "_cover.pdf"
LOG_XLSX_NAME: Final[str] = "_packet_log.xlsx"

PDF_EXTS: Final[frozenset[str]] = frozenset({".pdf"})
IMAGE_EXTS: Final[frozenset[str]] = frozenset({
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp",
    ".svg", ".heic", ".heif",
})


# ---------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------
def _log(prefix: str, msg: str) -> None:
    print(f"[{prefix}] {msg}", flush=True)


# ---------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------
@dataclass(slots=True)
class ItemLog:
    order: int
    file_type: str
    input_name: str
    input_path: str
    status: str
    output_pdf: str
    pages_added: int
    duration_sec: float
    message: str


@dataclass(slots=True)
class PackageResult:
    wi_id: int
    output_pdf: Path
    n_items: int
    n_failed: int
    n_pages: int
    log_xlsx: Path
    items: list[ItemLog] = field(default_factory=list)


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def _safe(s: str) -> str:
    bad = '<>:"/\\|?*'
    return "".join("_" if c in bad else c for c in s).strip(". ") or "x"


def _is_temp(p: Path) -> bool:
    return p.name.startswith("~$") or p.name.startswith("~")


def _list_attachments(
    folder: Path,
) -> tuple[list[Path], list[Path], list[Path]]:
    """Return (office_files, pdfs, images) in name-asc order."""
    office: list[Path] = []
    pdfs: list[Path] = []
    images: list[Path] = []
    if not folder.exists():
        return office, pdfs, images
    for p in folder.iterdir():
        if not p.is_file() or _is_temp(p):
            continue
        ext = p.suffix.lower()
        if ext in PDF_EXTS:
            pdfs.append(p.resolve())
        elif ext in IMAGE_EXTS:
            images.append(p.resolve())
        elif is_office_extension(ext):
            office.append(p.resolve())
    key = lambda x: x.name.lower()
    return (
        sorted(office, key=key),
        sorted(pdfs, key=key),
        sorted(images, key=key),
    )


def _pagesize(paper: str):
    return LETTER if paper.upper() == "LETTER" else A4


# ---------------------------------------------------------------------
# Cover page builder (preserves inline images)
# ---------------------------------------------------------------------
def _styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle(
            "WI_H1", parent=base["Heading1"], fontName="Helvetica-Bold",
            fontSize=18, leading=22, spaceAfter=12,
        ),
        "h2": ParagraphStyle(
            "WI_H2", parent=base["Heading2"], fontName="Helvetica-Bold",
            fontSize=13, leading=16, spaceBefore=12, spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "WI_Body", parent=base["BodyText"], fontName="Helvetica",
            fontSize=10, leading=14,
        ),
        "mono": ParagraphStyle(
            "WI_Mono", parent=base["BodyText"], fontName="Courier",
            fontSize=9, leading=12,
        ),
    }


def _safe_html(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _para_from_text(text: str, style: ParagraphStyle) -> list[Paragraph]:
    if not text.strip():
        return [Paragraph("(empty)", style)]
    paras: list[Paragraph] = []
    for chunk in text.split("\n\n"):
        chunk = chunk.strip()
        if not chunk:
            continue
        safe = _safe_html(chunk).replace("\n", "<br/>")
        paras.append(Paragraph(safe, style))
    return paras or [Paragraph("(empty)", style)]


_IMG_TAG_RE: Final[re.Pattern[str]] = re.compile(r"<img[^>]*?>", re.IGNORECASE)
_IMG_SRC_RE: Final[re.Pattern[str]] = re.compile(
    r"""src=["']([^"']+)["']""", re.IGNORECASE
)
_IMG_MARKER_RE: Final[re.Pattern[str]] = re.compile(r"__IMG_PLACEHOLDER_(\d+)__")


def _make_image_flowable(
    image_path: Path,
    max_width_cm: float = 14.0,
    max_height_cm: float = 18.0,
) -> Any:
    try:
        from PIL import Image as PILImage
        with PILImage.open(image_path) as img:
            w_px, h_px = img.size
        DPI = 96.0
        w_pt = w_px * 72.0 / DPI
        h_pt = h_px * 72.0 / DPI
        max_w = max_width_cm * cm
        max_h = max_height_cm * cm
        scale = min(1.0, max_w / max(w_pt, 0.1), max_h / max(h_pt, 0.1))
        w_pt *= scale
        h_pt *= scale
        return RLImage(str(image_path), width=w_pt, height=h_pt)
    except Exception as e:
        _log("ERROR", f"Image flowable failed for {image_path}: {e!r}")
        return None


def _html_to_flowables(
    html: str,
    inline_images_dir: Path,
    image_map: dict[str, str],
    body_style: ParagraphStyle,
) -> list:
    if not html or not html.strip():
        return [Paragraph("(empty)", body_style)]

    placeholders: list[str] = []

    def _replace(m: re.Match) -> str:
        sm = _IMG_SRC_RE.search(m.group(0))
        if not sm:
            return ""
        idx = len(placeholders)
        placeholders.append(sm.group(1))
        return f"\n__IMG_PLACEHOLDER_{idx}__\n"

    html_marked = _IMG_TAG_RE.sub(_replace, html)
    text = html_to_text(html_marked)

    parts = _IMG_MARKER_RE.split(text)
    flowables: list = []
    for i, part in enumerate(parts):
        if i % 2 == 0:
            if part.strip():
                flowables.extend(_para_from_text(part, body_style))
        else:
            idx = int(part)
            if idx < len(placeholders):
                url = placeholders[idx]
                filename = image_map.get(url)
                if filename:
                    local = inline_images_dir / filename
                    if local.exists():
                        flow = _make_image_flowable(local)
                        if flow is not None:
                            flowables.append(flow)
                            flowables.append(Spacer(1, 0.3 * cm))
                            continue
                flowables.append(Paragraph(
                    "<i>[inline image unavailable]</i>", body_style
                ))

    return flowables or [Paragraph("(empty)", body_style)]


def _render_comments_section(
    wi_dir: Path,
    inline_images_dir: Path,
    image_map: dict[str, str],
    sty: dict[str, ParagraphStyle],
) -> list:
    comments_json_path = wi_dir / "_comments.json"
    if comments_json_path.exists():
        try:
            comments = json.loads(comments_json_path.read_text(encoding="utf-8"))
        except Exception:
            comments = None
        if isinstance(comments, list):
            if not comments:
                return [Paragraph("(no comments)", sty["body"])]
            flowables: list = []
            for i, c in enumerate(comments):
                if i > 0:
                    flowables.append(Spacer(1, 0.3 * cm))
                header = _safe_html(
                    f"[{c.get('when', '')}] {c.get('author', 'unknown')}"
                )
                flowables.append(Paragraph(header, sty["mono"]))
                flowables.extend(_html_to_flowables(
                    c.get("html", ""), inline_images_dir, image_map, sty["body"]
                ))
            return flowables

    com_path = wi_dir / "_comments.txt"
    com = com_path.read_text(encoding="utf-8") if com_path.exists() else ""
    return _para_from_text(com, sty["body"])


def make_cover_pdf(wi_dir: Path, out_pdf: Path, paper: str) -> None:
    meta_path = wi_dir / "_meta.json"
    desc_html_path = wi_dir / "_description.html"
    ac_html_path = wi_dir / "_ac.html"
    image_map_path = wi_dir / "_inline_images.json"
    inline_images_dir = wi_dir / "inline_images"

    meta: dict = {}
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}

    desc_html = (
        desc_html_path.read_text(encoding="utf-8")
        if desc_html_path.exists() else ""
    )
    ac_html = (
        ac_html_path.read_text(encoding="utf-8")
        if ac_html_path.exists() else ""
    )

    image_map: dict[str, str] = {}
    if image_map_path.exists():
        try:
            image_map = json.loads(image_map_path.read_text(encoding="utf-8"))
        except Exception:
            image_map = {}

    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_pdf), pagesize=_pagesize(paper),
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"WI {meta.get('wi_id', '')} cover",
    )
    sty = _styles()

    story: list = []
    wi_id_str = str(meta.get("wi_id", "?"))
    wi_url = meta.get("wi_url", "")
    wi_title = _safe_html(meta.get("title", "(no title)"))
    if wi_url:
        heading = f'<a href="{_safe_html(wi_url)}" color="blue">WI {wi_id_str}</a>: {wi_title}'
    else:
        heading = f"WI {wi_id_str}: {wi_title}"
    story.append(Paragraph(heading, sty["h1"]))
    meta_lines = [
        _safe_html(f"Type: {meta.get('type', '')}"),
        _safe_html(f"State: {meta.get('state', '')}"),
        _safe_html(f"Board Lane: {meta.get('board_lane', '')}"),
        _safe_html(f"Iteration: {meta.get('iteration', '')}"),
        _safe_html(f"Area: {meta.get('area', '')}"),
    ]
    story.append(Paragraph("<br/>".join(meta_lines), sty["body"]))

    story.append(Paragraph("Description", sty["h2"]))
    story.extend(_html_to_flowables(
        desc_html, inline_images_dir, image_map, sty["body"]
    ))

    story.append(Paragraph("Acceptance Criteria", sty["h2"]))
    story.extend(_html_to_flowables(
        ac_html, inline_images_dir, image_map, sty["body"]
    ))

    story.append(PageBreak())
    story.append(Paragraph("Comments", sty["h2"]))
    story.extend(_render_comments_section(
        wi_dir, inline_images_dir, image_map, sty
    ))

    doc.build(story)


def _make_attachments_separator_pdf(
    out_pdf: Path, n_attachments: int, paper: str,
) -> None:
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_pdf), pagesize=_pagesize(paper),
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Attachments",
    )
    sty = _styles()
    story = [
        Spacer(1, 6 * cm),
        Paragraph("Attachments", sty["h1"]),
        Spacer(1, 0.5 * cm),
        Paragraph(
            f"The following pages contain {n_attachments} attachment file(s) "
            "in alphabetical order.",
            sty["body"],
        ),
    ]
    doc.build(story)


# ---------------------------------------------------------------------
# Error placeholder
# ---------------------------------------------------------------------
def _placeholder_pdf(
    out_pdf: Path, title: str, details: str, paper: str = "A4",
) -> None:
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(out_pdf), pagesize=_pagesize(paper))
    c.setTitle(title[:120])
    t = c.beginText(40, 800)
    t.setLeading(14)
    t.textLine("ERROR PLACEHOLDER (Conversion Failed)")
    t.textLine("")
    t.textLine(f"Item: {title}")
    t.textLine("")
    for line in (details or "(no details)").splitlines():
        line = line.strip()
        while len(line) > 110:
            t.textLine(line[:110])
            line = line[110:]
        t.textLine(line)
    c.drawText(t)
    c.showPage()
    c.save()


# ---------------------------------------------------------------------
# Image -> PDF
# ---------------------------------------------------------------------
def _image_to_pdf(img_path: Path, out_pdf: Path) -> None:
    from PIL import Image
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    ext = img_path.suffix.lower()

    # SVG: attempt cairosvg conversion, fall back to placeholder
    if ext == ".svg":
        try:
            import cairosvg  # type: ignore
            cairosvg.svg2pdf(url=str(img_path), write_to=str(out_pdf))
            return
        except ImportError:
            raise RuntimeError(
                "SVG rendering requires cairosvg (pip install cairosvg). "
                f"File: {img_path.name}"
            )

    # HEIC/HEIF: attempt pillow-heif registration
    if ext in (".heic", ".heif"):
        try:
            import pillow_heif  # type: ignore
            pillow_heif.register_heif_opener()
        except ImportError:
            raise RuntimeError(
                "HEIC/HEIF support requires pillow-heif "
                "(pip install pillow-heif). "
                f"File: {img_path.name}"
            )

    with Image.open(img_path) as im:
        if im.mode in ("RGBA", "LA", "P", "CMYK", "YCbCr", "LAB", "HSV",
                       "I", "F"):
            im = im.convert("RGB")
        # Constrain very large images to avoid memory exhaustion
        max_dim = 4096
        if im.width > max_dim or im.height > max_dim:
            im.thumbnail((max_dim, max_dim), Image.LANCZOS)
        im.save(out_pdf, "PDF", resolution=150.0)


# ---------------------------------------------------------------------
# Final merge
# ---------------------------------------------------------------------
def _merge_strict(ordered: list[Path], out_pdf: Path) -> list[int]:
    writer = PdfWriter()
    pages_per: list[int] = []
    for p in ordered:
        try:
            reader = PdfReader(str(p))
            # Attempt to decrypt if the PDF is encrypted (try empty password
            # first which handles owner-only restricted PDFs)
            if reader.is_encrypted:
                try:
                    reader.decrypt("")
                except Exception:
                    raise RuntimeError(
                        f"Encrypted PDF (password required): {p.name}"
                    )
            n = len(reader.pages)
            for page in reader.pages:
                writer.add_page(page)
            pages_per.append(n)
        except Exception as e:
            ph = p.parent / f"__UNREADABLE_{_safe(p.stem)}.pdf"
            _placeholder_pdf(ph, p.name, f"Unreadable PDF: {e!r}\nPath: {p}")
            reader = PdfReader(str(ph))
            n = len(reader.pages)
            for page in reader.pages:
                writer.add_page(page)
            pages_per.append(n)
    out_pdf.parent.mkdir(parents=True, exist_ok=True)
    with out_pdf.open("wb") as f:
        writer.write(f)
    return pages_per


# ---------------------------------------------------------------------
# Log writer
# ---------------------------------------------------------------------
def _write_log_xlsx(items: list[ItemLog], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append([
        "Order", "Type", "Input Name", "Input Path", "Status",
        "Output PDF", "Pages Added", "Duration (sec)", "Message",
    ])
    for r in items:
        ws.append([
            r.order, r.file_type, r.input_name, r.input_path, r.status,
            r.output_pdf, r.pages_added, r.duration_sec, r.message,
        ])
    widths = [6, 8, 35, 60, 10, 60, 12, 14, 90]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(out_path)


# =====================================================================
# Public entry point
# =====================================================================
def package_for_wi(
    wi_dir: Path,
    paper_size: str = "A4",
    output_pdf: Path | None = None,
) -> PackageResult:
    wi_dir = wi_dir.resolve()
    meta_path = wi_dir / "_meta.json"
    if not meta_path.exists():
        raise FileNotFoundError(f"_meta.json missing in {wi_dir}")
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    wi_id = int(meta.get("wi_id", 0))

    out_dir = wi_dir
    if output_pdf is None:
        output_pdf = out_dir / PACKET_PDF_NAME.format(wi_id=wi_id)
    output_pdf.parent.mkdir(parents=True, exist_ok=True)

    cover = wi_dir / COVER_PDF_NAME
    make_cover_pdf(wi_dir, cover, paper_size)

    att_dir = wi_dir / "attachments"
    office_files, pdfs, images = _list_attachments(att_dir)

    items: list[ItemLog] = []
    ordered: list[Path] = [cover]
    items.append(ItemLog(
        order=0, file_type="COVER", input_name=cover.name,
        input_path=str(cover), status="SUCCESS",
        output_pdf=str(cover), pages_added=0, duration_sec=0.0,
        message="Cover from ADO text fields",
    ))

    order = 1
    total_attachments = len(office_files) + len(pdfs) + len(images)
    separator: Path | None = None
    if total_attachments > 0:
        separator = wi_dir / "_attachments_separator.pdf"
        _make_attachments_separator_pdf(separator, total_attachments, paper_size)
        items.append(ItemLog(
            order=order, file_type="HEADER", input_name=separator.name,
            input_path=str(separator), status="SUCCESS",
            output_pdf=str(separator), pages_added=0, duration_sec=0.0,
            message="Attachments section header",
        ))
        ordered.append(separator)
        order += 1

    # Office files (xlsx, docx, csv, rtf, txt) - via pure-Python converter
    for src in office_files:
        t0 = time.time()
        out = out_dir / f"_att_{_safe(src.stem)}.pdf"
        ext = src.suffix.lower()
        ftype = "EXCEL" if ext in {
            ".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".xlsb",
            ".csv", ".tsv",
        } else "WORD" if ext in {
            ".docx", ".docm", ".dotx", ".dotm", ".doc", ".dot", ".rtf",
        } else "TEXT"
        try:
            status, msg = convert_to_pdf(src, out, paper_size)
        except Exception as e:
            status, msg = "FAILED", f"Converter exception: {e!r}"
        if not out.exists() or out.stat().st_size == 0 or status == "FAILED":
            ph = out_dir / f"__FAILED__{_safe(src.stem)}.pdf"
            _placeholder_pdf(
                ph, src.name,
                f"Conversion failed.\n{msg}\nPath: {src}",
                paper_size,
            )
            used = ph
        else:
            used = out
        items.append(ItemLog(
            order=order, file_type=ftype, input_name=src.name,
            input_path=str(src), status=status, output_pdf=str(used),
            pages_added=0, duration_sec=round(time.time() - t0, 3), message=msg,
        ))
        ordered.append(used)
        order += 1

    # PDFs - passthrough
    for src in pdfs:
        items.append(ItemLog(
            order=order, file_type="PDF", input_name=src.name,
            input_path=str(src), status="SUCCESS", output_pdf=str(src),
            pages_added=0, duration_sec=0.0, message="Passthrough",
        ))
        ordered.append(src)
        order += 1

    # Images
    for src in images:
        t0 = time.time()
        out = out_dir / f"_att_{_safe(src.stem)}.pdf"
        try:
            _image_to_pdf(src, out)
            status, msg, used = "SUCCESS", "Image converted", out
        except Exception as e:
            ph = out_dir / f"__FAILED__{_safe(src.stem)}.pdf"
            _placeholder_pdf(
                ph, src.name, f"Image failed: {e!r}", paper_size,
            )
            status, msg, used = "FAILED", f"Image exception: {e!r}", ph
        items.append(ItemLog(
            order=order, file_type="IMAGE", input_name=src.name,
            input_path=str(src), status=status, output_pdf=str(used),
            pages_added=0, duration_sec=round(time.time() - t0, 3), message=msg,
        ))
        ordered.append(used)
        order += 1

    pages_per = _merge_strict(ordered, output_pdf)
    for i, n in enumerate(pages_per):
        items[i].pages_added = n

    # Cleanup intermediate files
    for log in items:
        p = Path(log.output_pdf)
        if p.exists() and p != cover and p != output_pdf and p.parent == out_dir:
            if separator is not None and p == separator:
                continue
            try:
                p.unlink()
            except Exception:
                pass
    try:
        cover.unlink()
    except Exception:
        pass
    if separator is not None:
        try:
            separator.unlink()
        except Exception:
            pass

    log_xlsx = out_dir / LOG_XLSX_NAME
    _write_log_xlsx(items, log_xlsx)

    n_failed = sum(1 for it in items if it.status == "FAILED")
    _log(
        "INFO",
        f"WI {wi_id} packet -> {output_pdf.name} | items={len(items)} "
        f"failed={n_failed} pages={sum(pages_per)}",
    )

    result = PackageResult(
        wi_id=wi_id,
        output_pdf=output_pdf,
        n_items=len(items),
        n_failed=n_failed,
        n_pages=sum(pages_per),
        log_xlsx=log_xlsx,
        items=items,
    )

    del ordered, items, pages_per
    gc.collect()
    return result
