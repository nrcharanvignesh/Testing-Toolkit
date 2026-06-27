"""
template_analyzer.py
LLM-assisted template structure analysis for client Excel test-script templates.

Templates uploaded by users are typically NOT empty — they contain existing
test data for a different scenario. This module performs a deep structural
analysis of the workbook (sheet organization, formatting, styling, merged
cells, column widths, frozen panes, data patterns) and sends a comprehensive
description to the LLM. The LLM identifies the layout structure: which row
is the header, which columns map to which test-case fields, and what the
row organization pattern is (one row per step vs one row per test case).

Once analyzed, the resulting spec is deterministic — the same template always
produces the same output. The LLM is only involved at upload time; all
subsequent renders use the cached spec without any AI calls.

Public API:
    analyze_template_with_llm(client, model, xlsx_path) -> (header_row, mapping)
"""

from __future__ import annotations

import gc
import json
import re
from pathlib import Path
from typing import Any, Final

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from testgen.testcase_template import FIELDS, _pick_sheet

_MAX_SCAN_ROWS: Final[int] = 50

_SYSTEM_PROMPT: Final[str] = """\
You are an expert at analyzing Excel test-case/test-script templates used by \
QA teams in enterprise software projects.

You are given a DETAILED structural description of an Excel workbook: its \
sheets, formatting, column widths, merged cells, frozen panes, and the first \
~50 rows of data. The workbook likely contains EXISTING test data for a \
different scenario — your job is to understand the STRUCTURE and LAYOUT, not \
the specific data content.

CRITICAL: Your analysis MUST be thorough and exact. The output of this \
analysis will be used to generate test scripts that must PERFECTLY replicate \
the template's layout. Every column must be correctly identified. Missing or \
incorrect mappings will cause output quality degradation.

Your task:
1. Identify which row is the HEADER row (the row with column titles/labels). \
Look for the row with the most non-empty cells that appear to be labels \
rather than data. Headers are often bold, in the row just above the first \
data row, or just below frozen panes.
2. Determine the row organization: "step" mode (one row per test step, with \
test-case-level fields repeating or merged on the first step) or "tc" mode \
(one row per test case, with steps joined into multi-line cells).
3. Map EVERY relevant column to the appropriate test-case field based on the \
header text AND the data patterns you observe. Do NOT skip columns unless \
they are purely decorative (borders, spacers, serial numbers unrelated to \
test cases).

Valid field keys (use these exactly):
- test_case_id: sequence/ID column for the test case number
- story_id: parent work item / user story ID / requirement ID
- story_title: parent work item title or requirement description
- tc_index: index of the test case within its parent story
- tc_title: test case name, title, scenario, objective, or description
- category: test type, module, or category
- priority: priority or severity level
- tags: labels, tags, or keywords
- preconditions: preconditions, prerequisites, setup steps, or assumptions
- step_index: step number or sequence within a test case
- action: the test step action, procedure, instruction, or "steps to execute"
- expected: the expected result, expected outcome, or acceptance criteria
- comments: remarks, notes, actual result, status, or observation column

Analysis guidance:
- Look at DATA PATTERNS to confirm header meanings. If a column header says \
"Description" but the data contains step-by-step actions, map it to "action".
- If headers are in a non-English language, map by semantic meaning.
- If you see merged cells spanning multiple rows in a column (like test case \
title spanning its steps), that confirms "step" mode.
- If a column contains numbered lists within cells ("1. Do X\\n2. Do Y"), \
that confirms "tc" mode for that field.
- Examine the DATA in rows below the header carefully. The actual content \
tells you what a column is for, even if the header is ambiguous.
- If multiple columns could map to the same field, pick the one with the \
most relevant data patterns.
- Column numbers are 1-based (A=1, B=2, C=3, etc.).
- Map ALL columns that have clear test-case semantics. Err on the side of \
including a mapping rather than skipping it.

Return ONLY valid JSON (no markdown fences, no explanation):
{
  "header_row": <1-based row number>,
  "row_mode": "step" or "tc",
  "mapping": {"field_name": column_number, ...}
}\
"""


def _extract_full_structure(xlsx_path: Path) -> tuple[str, str]:
    """Extract comprehensive structural info from the workbook.

    Returns (sheet_name, description_text) where description_text is a
    rich representation of the sheet's structure for LLM analysis."""
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=False)
    try:
        sheet_name = _pick_sheet(wb)
        ws = wb[sheet_name]
        max_col = min(ws.max_column or 1, 26)
        max_row = min(ws.max_row or 1, _MAX_SCAN_ROWS)

        parts: list[str] = []

        parts.append(f"WORKBOOK SHEETS: {', '.join(wb.sheetnames)}")
        parts.append(f"ACTIVE SHEET: '{sheet_name}'")
        parts.append(f"DIMENSIONS: {ws.dimensions} ({max_row}+ rows, "
                     f"{max_col} columns)")

        if ws.freeze_panes:
            parts.append(f"FROZEN PANES: {ws.freeze_panes}")

        col_widths: list[str] = []
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            dim = ws.column_dimensions.get(letter)
            w = dim.width if dim and dim.width else "default"
            col_widths.append(f"{letter}={w}")
        parts.append(f"COLUMN WIDTHS: {', '.join(col_widths)}")

        merged: list[str] = []
        for rng in ws.merged_cells.ranges:
            merged.append(str(rng))
        if merged:
            parts.append(f"MERGED CELLS: {', '.join(merged[:20])}")
            if len(merged) > 20:
                parts.append(f"  ... and {len(merged) - 20} more")

        parts.append("")
        parts.append("DATA GRID (first rows):")
        parts.append("")

        header_line = "      " + " | ".join(
            f"{get_column_letter(c):^12}" for c in range(1, max_col + 1)
        )
        parts.append(header_line)
        parts.append("      " + "-" * (max_col * 15))

        for r in range(1, max_row + 1):
            cells: list[str] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                text = "" if v is None else str(v).strip()
                if len(text) > 60:
                    text = text[:57] + "..."
                bold = ""
                if hasattr(cell, "font") and cell.font and cell.font.bold:
                    bold = "[B]"
                cells.append(f"{bold}{text:^14}"[:14])
            parts.append(f"R{r:>4} " + " | ".join(cells))

        return sheet_name, "\n".join(parts)
    finally:
        wb.close()
        del wb
        gc.collect()


def _build_user_prompt(sheet_name: str, structure: str) -> str:
    return (
        f"Analyze this Excel test-script template. It contains EXISTING "
        f"test data for another scenario — focus on understanding the "
        f"STRUCTURE, LAYOUT, and COLUMN PURPOSES, not the specific data.\n\n"
        f"{structure}\n\n"
        f"Based on the structure, formatting, headers, and data patterns, "
        f"identify the header row, row mode, and column-to-field mapping."
    )


def _parse_response(text: str) -> tuple[int | None, str | None, dict[str, int]]:
    """Extract header_row, row_mode, and mapping from LLM response."""
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    data = json.loads(cleaned)
    header_row = data.get("header_row")
    if not (isinstance(header_row, int) and header_row >= 1):
        header_row = None

    row_mode = data.get("row_mode")
    if row_mode not in ("step", "tc"):
        row_mode = None

    raw_mapping = data.get("mapping", {})
    mapping: dict[str, int] = {}
    for fld, col in raw_mapping.items():
        if fld in FIELDS and isinstance(col, int) and col >= 1:
            mapping[fld] = col

    return header_row, row_mode, mapping


def analyze_template_with_llm(
    client: Any,
    model: str,
    xlsx_path: Path | str,
) -> tuple[int | None, dict[str, int]]:
    """Send the template's full structure to the LLM for semantic analysis.

    The template likely contains existing data for a different scenario.
    The LLM analyzes structure, formatting, styling, sheet organization,
    and data patterns to determine the correct field-to-column mapping.

    Returns (header_row_override, field_to_column_mapping).
    header_row_override is None if the LLM didn't determine one.
    The mapping dict maps field names to 1-based column numbers.

    Once this analysis is saved as a spec, all subsequent renders are
    fully deterministic with no further LLM calls.

    Uses extended thinking when available for maximum accuracy.

    Raises on API or parse failure (caller should catch and fall back)."""
    xlsx_path = Path(xlsx_path)
    _sheet_name, structure = _extract_full_structure(xlsx_path)
    user_msg = _build_user_prompt(_sheet_name, structure)

    # Use extended thinking for thorough analysis when model supports it
    thinking_budget = 4000
    try:
        result = client.complete(
            model=model,
            system=_SYSTEM_PROMPT,
            user=user_msg,
            max_tokens=2500,
            temperature=0.0,
            thinking_budget=thinking_budget,
        )
    except (TypeError, Exception):
        # Fallback without thinking if not supported
        result = client.complete(
            model=model,
            system=_SYSTEM_PROMPT,
            user=user_msg,
            max_tokens=2500,
            temperature=0.0,
        )

    header_row, row_mode, mapping = _parse_response(result.text)

    # Validate mapping quality: must have at least action OR tc_title
    if not mapping or not (mapping.get("action") or mapping.get("tc_title")):
        # Retry once with explicit instruction to map more columns
        retry_msg = (
            f"{user_msg}\n\nIMPORTANT: Your previous analysis found "
            f"insufficient mappings. You MUST map at minimum: tc_title "
            f"and action columns. Look at the data patterns more carefully."
        )
        result = client.complete(
            model=model,
            system=_SYSTEM_PROMPT,
            user=retry_msg,
            max_tokens=2500,
            temperature=0.0,
        )
        header_row2, row_mode2, mapping2 = _parse_response(result.text)
        if mapping2 and len(mapping2) > len(mapping):
            header_row, row_mode, mapping = header_row2, row_mode2, mapping2

    _ = row_mode  # reserved for future use if spec gains explicit row_mode
    return header_row, mapping
