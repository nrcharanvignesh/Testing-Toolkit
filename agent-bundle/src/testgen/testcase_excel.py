"""
testcase_excel.py
Round-trip between the ADO TC JSON payload and an Excel review file.

Workflow:
    LLM JSON  --payload_to_xlsx-->  review.xlsx  (human edits in Excel)
    review.xlsx  --xlsx_to_payload-->  edited JSON  -->  create_test_cases()

Excel layout:
    One row per test STEP. TC-level fields (title, category, priority,
    tags, preconditions) are written on the row of the first step of
    that TC and left blank on subsequent step rows for that same TC.
    On read-back we forward-fill within each TC group.

    Columns (in order):
        A  Story #         parent_work_item_id (int)
        B  Story Title     parent_title (str, reference)
        C  TC #            1-based index within the story (int)
        D  TC Title        test_case.title (str)
        E  Category        one of VALID_CATEGORIES (dropdown)
        F  Priority        one of VALID_PRIORITIES (dropdown)
        G  Tags            "; "-joined tag list (str)
        H  Preconditions   str
        I  Skip            "Yes" / "No" (dropdown; default "No")
        J  Step #          1-based step index (int)
        K  Action          str
        L  Expected        str

    A frozen header row, column auto-widths, and data validation
    dropdowns are applied so the file is friendly to edit in Excel.

NOTES
-----
- Step boundaries are determined by Step # restarting at 1 (or by an
  empty TC Title cell beneath a populated one).
- A "Yes" in the Skip column drops the entire TC from the output.
- schema_version is hard-coded to 1 on write and ignored on read.
"""

from __future__ import annotations

import gc
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from ado.testcase_creator import (
    VALID_CATEGORIES,
    VALID_PRIORITIES,
    clean_title,
)

SHEET_NAME = "Test Cases"
HEADERS: list[str] = [
    "Story #", "Story Title", "TC #", "TC Title", "Category", "Priority",
    "Tags", "Preconditions", "Skip", "Step #", "Action", "Expected",
]
# Approximate widths in Excel character units
COL_WIDTHS: dict[str, int] = {
    "A": 10, "B": 36, "C": 6,  "D": 60, "E": 18, "F": 10,
    "G": 24, "H": 36, "I": 7,  "J": 7,  "K": 60, "L": 60,
}

# Column index helpers (1-based for openpyxl)
COL_STORY_ID    = 1
COL_STORY_TITLE = 2
COL_TC_INDEX    = 3
COL_TC_TITLE    = 4
COL_CATEGORY    = 5
COL_PRIORITY    = 6
COL_TAGS        = 7
COL_PRECONDITIONS = 8
COL_SKIP        = 9
COL_STEP_INDEX  = 10
COL_ACTION      = 11
COL_EXPECTED    = 12


# ---------------------------------------------------------------------
# Write: payload -> xlsx
# ---------------------------------------------------------------------
def payload_to_xlsx(payload: dict[str, Any], out_path: Path) -> int:
    """Write the JSON payload to an Excel review file. Returns the
    total number of step rows written."""
    out_path = Path(out_path)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _write_header(ws)

    row = 2
    n_rows_written = 0
    stories = payload.get("stories") or []
    for story in stories:
        parent_id = story.get("parent_work_item_id")
        parent_title = (story.get("parent_title") or "").strip()
        tcs = story.get("test_cases") or []
        for tc_idx, tc in enumerate(tcs, start=1):
            steps = tc.get("steps") or [{"action": "", "expected": ""}]
            title = clean_title((tc.get("title") or "").strip())
            category = tc.get("category") or ""
            priority = tc.get("priority") or ""
            tags = "; ".join(
                str(t) for t in (tc.get("tags") or []) if t
            )
            pre = (tc.get("preconditions") or "").strip()
            for step_idx, st in enumerate(steps, start=1):
                _write_cell(ws, row, COL_STORY_ID, parent_id)
                _write_cell(ws, row, COL_STORY_TITLE, parent_title)
                _write_cell(ws, row, COL_TC_INDEX, tc_idx)
                # TC-level fields only on the first step row
                if step_idx == 1:
                    _write_cell(ws, row, COL_TC_TITLE, title)
                    _write_cell(ws, row, COL_CATEGORY, category)
                    _write_cell(ws, row, COL_PRIORITY, priority)
                    _write_cell(ws, row, COL_TAGS, tags)
                    _write_cell(ws, row, COL_PRECONDITIONS, pre)
                    _write_cell(ws, row, COL_SKIP, "No")
                _write_cell(ws, row, COL_STEP_INDEX, step_idx)
                _write_cell(ws, row, COL_ACTION, st.get("action", ""))
                _write_cell(ws, row, COL_EXPECTED, st.get("expected", ""))
                row += 1
                n_rows_written += 1

    _apply_widths(ws)
    _apply_validations(ws, last_row=row - 1)
    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    wb.close()
    del wb
    gc.collect()
    return n_rows_written


def _write_header(ws: Worksheet) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5B8A")
    align = Alignment(vertical="center", horizontal="left", wrap_text=False)
    for col_idx, label in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align
    ws.row_dimensions[1].height = 22


def _write_cell(ws: Worksheet, row: int, col: int, value: Any) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(vertical="top", wrap_text=True)


def _apply_widths(ws: Worksheet) -> None:
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _apply_validations(ws: Worksheet, last_row: int) -> None:
    """Attach dropdown validations to Category, Priority, and Skip
    columns for rows 2..last_row. Excel limits in-cell list strings
    to 255 chars; we are well under that."""
    if last_row < 2:
        return

    def _add_dropdown(values: tuple[str, ...], col_letter: str) -> None:
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Pick one of the listed values."
        dv.errorTitle = "Invalid value"
        dv.prompt = "Choose from the dropdown."
        dv.promptTitle = "Allowed values"
        rng = f"{col_letter}2:{col_letter}{last_row}"
        dv.add(rng)
        ws.add_data_validation(dv)

    _add_dropdown(VALID_CATEGORIES, get_column_letter(COL_CATEGORY))
    _add_dropdown(VALID_PRIORITIES, get_column_letter(COL_PRIORITY))
    _add_dropdown(("Yes", "No"),    get_column_letter(COL_SKIP))


# ---------------------------------------------------------------------
# Read: xlsx -> payload
# ---------------------------------------------------------------------
class ExcelParseError(ValueError):
    """Raised when the review Excel can't be reconstructed into a
    valid payload."""


def xlsx_to_payload(in_path: Path) -> tuple[dict[str, Any], list[str]]:
    """Read a review Excel back into the JSON payload format. Returns
    (payload, warnings). Raises ExcelParseError on structural problems.

    Forward-fill semantics: TC-level fields (title, category, priority,
    tags, preconditions, skip) are inherited from the most recent row
    where they were populated within the same (story, tc_index) group.
    """
    in_path = Path(in_path)
    wb = load_workbook(filename=str(in_path), data_only=True, read_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        # Tolerate renames; use first sheet
        ws = wb[wb.sheetnames[0]]

    warnings: list[str] = []
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        wb.close()
        raise ExcelParseError("Workbook is empty.")
    header = [str(v or "").strip() for v in rows[0]]
    expected_min = ["Story #", "TC #", "TC Title", "Step #", "Action", "Expected"]
    missing = [h for h in expected_min if h not in header]
    if missing:
        wb.close()
        raise ExcelParseError(
            f"Header row missing required columns: {missing}. "
            f"Header found: {header}"
        )

    # Map header name to column index for tolerance to reordering
    idx = {h: i for i, h in enumerate(header) if h}

    def col(name: str, row_tuple: tuple, default: Any = "") -> Any:
        i = idx.get(name)
        if i is None or i >= len(row_tuple):
            return default
        v = row_tuple[i]
        return v if v is not None else default

    # Group by (story_id, tc_index_within_story). Each group becomes a
    # single test case object.
    Group = dict[str, Any]  # type alias for readability
    groups: dict[tuple[int, int], Group] = {}
    group_order: list[tuple[int, int]] = []
    story_titles: dict[int, str] = {}

    last_tc_meta: dict[tuple[int, int], dict[str, Any]] = {}

    for row_num, row in enumerate(rows[1:], start=2):
        # Skip entirely blank rows
        if not any(v not in (None, "") for v in row):
            continue

        story_id_raw = col("Story #", row)
        try:
            story_id = int(story_id_raw)
        except (TypeError, ValueError):
            warnings.append(
                f"Row {row_num}: Story # = {story_id_raw!r} is not an int; skipping row."
            )
            continue

        tc_idx_raw = col("TC #", row)
        try:
            tc_idx = int(tc_idx_raw)
        except (TypeError, ValueError):
            warnings.append(
                f"Row {row_num}: TC # = {tc_idx_raw!r} is not an int; skipping row."
            )
            continue

        st_idx_raw = col("Step #", row)
        try:
            st_idx = int(st_idx_raw)
        except (TypeError, ValueError):
            warnings.append(
                f"Row {row_num}: Step # = {st_idx_raw!r} is not an int; skipping row."
            )
            continue

        key = (story_id, tc_idx)
        if key not in groups:
            groups[key] = {
                "title": "",
                "category": "",
                "priority": None,
                "tags": [],
                "preconditions": "",
                "skip": False,
                "steps": [],
            }
            group_order.append(key)

        # Story title: latest non-empty wins
        story_title_v = str(col("Story Title", row) or "").strip()
        if story_title_v:
            story_titles[story_id] = story_title_v
        elif story_id not in story_titles:
            story_titles[story_id] = ""

        # TC-level: forward-fill within this group
        g = groups[key]
        tc_title = str(col("TC Title", row) or "").strip()
        if tc_title:
            g["title"] = clean_title(tc_title)
        cat = str(col("Category", row) or "").strip()
        if cat:
            g["category"] = cat
        pri = str(col("Priority", row) or "").strip()
        if pri:
            g["priority"] = pri
        tags_raw = str(col("Tags", row) or "").strip()
        if tags_raw:
            g["tags"] = [
                t.strip().lower() for t in tags_raw.split(";") if t.strip()
            ]
        pre = str(col("Preconditions", row) or "").strip()
        if pre:
            g["preconditions"] = pre
        skip_raw = str(col("Skip", row) or "").strip().lower()
        if skip_raw in ("yes", "y", "true", "1"):
            g["skip"] = True

        # Step itself
        action = str(col("Action", row) or "").strip()
        expected = str(col("Expected", row) or "").strip()
        if not action and not expected:
            # No content on this row - skip silently. This catches rows
            # where the reviewer left only TC metadata without a step.
            continue
        g["steps"].append({
            "action": action, "expected": expected, "_idx": st_idx,
        })

    wb.close()
    del wb
    gc.collect()

    if not groups:
        raise ExcelParseError(
            "No usable data rows found. Ensure each row has at least "
            "Story #, TC #, Step #, and either Action or Expected text."
        )

    # Reorganize by story preserving first-seen order
    payload: dict[str, Any] = {"schema_version": 1, "stories": []}
    seen_story_order: list[int] = []
    by_story: dict[int, list[tuple[int, dict[str, Any]]]] = {}
    for key in group_order:
        sid, tc_idx = key
        if sid not in by_story:
            by_story[sid] = []
            seen_story_order.append(sid)
        by_story[sid].append((tc_idx, groups[key]))

    n_skipped = 0
    for sid in seen_story_order:
        tcs_out: list[dict[str, Any]] = []
        for _tc_idx, g in sorted(by_story[sid], key=lambda x: x[0]):
            if g["skip"]:
                n_skipped += 1
                continue
            if not g["steps"]:
                warnings.append(
                    f"Story {sid} TC '{g['title'] or '(no title)'}': "
                    f"no step rows; dropping."
                )
                continue
            # Sort steps by their own index, drop the synthetic _idx
            steps_sorted = sorted(g["steps"], key=lambda s: s.get("_idx", 0))
            steps_clean = [
                {"action": s["action"], "expected": s["expected"]}
                for s in steps_sorted
            ]
            tc_out: dict[str, Any] = {
                "title": g["title"],
                "category": g["category"],
                "steps": steps_clean,
            }
            if g["priority"]:
                tc_out["priority"] = g["priority"]
            if g["tags"]:
                tc_out["tags"] = g["tags"]
            if g["preconditions"]:
                tc_out["preconditions"] = g["preconditions"]
            # Constant custom_fields - the agent always emits these,
            # so preserve them on round-trip
            tc_out["custom_fields"] = {
                "QA GenAI Automated": "None",
                "QA GenAI Tool":      "None",
            }
            tcs_out.append(tc_out)
        if tcs_out:
            payload["stories"].append({
                "parent_work_item_id": sid,
                "parent_title": story_titles.get(sid, ""),
                "test_cases": tcs_out,
            })

    if n_skipped:
        warnings.append(
            f"Skipped {n_skipped} test case(s) marked Skip=Yes."
        )
    return payload, warnings
