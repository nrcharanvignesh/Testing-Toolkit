"""
defects/review_excel.py
Generate a review Excel from parsed defects (with embedded images) and
read back the reviewed version for ADO upload.
"""

from __future__ import annotations

import base64
import io
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from defects.parser import ParsedDefect

_HEADERS = [
    "Parent WI ID", "Title", "Description", "Repro Steps",
    "Severity", "Expected Result", "Actual Result", "Skip", "Images",
]
_SEVERITY_VALUES = ("Critical", "High", "Medium", "Low")
_HEADER_FILL = PatternFill("solid", fgColor="2B579A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def defects_to_xlsx(
    defects: list[ParsedDefect], out_path: Path | str,
) -> int:
    """Write defects to a review Excel file. Returns the number of rows."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defects Review"

    # Header row.
    for col, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths.
    widths = [12, 40, 50, 50, 12, 35, 35, 6, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows.
    for row_idx, defect in enumerate(defects, 2):
        ws.cell(row=row_idx, column=1, value=defect.parent_id or "")
        ws.cell(row=row_idx, column=2, value=defect.title)
        ws.cell(row=row_idx, column=3, value=defect.description)
        ws.cell(row=row_idx, column=4, value=defect.repro_steps)
        ws.cell(row=row_idx, column=5, value=defect.severity or "Medium")
        ws.cell(row=row_idx, column=6, value=defect.expected_result)
        ws.cell(row=row_idx, column=7, value=defect.actual_result)
        ws.cell(row=row_idx, column=8, value="No")

        # Embed first image (if any) in the Images column.
        if defect.images:
            try:
                img_data = base64.b64decode(defect.images[0].data_b64)
                img_stream = io.BytesIO(img_data)
                img = XlImage(img_stream)
                img.width = 100
                img.height = 75
                anchor = f"{get_column_letter(9)}{row_idx}"
                ws.add_image(img, anchor)
                ws.row_dimensions[row_idx].height = 60
            except Exception:
                ws.cell(row=row_idx, column=9,
                        value=f"({len(defect.images)} image(s))")
        else:
            ws.cell(row=row_idx, column=9, value="")

        # Wrap text for long fields.
        for col in (3, 4, 6, 7):
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    wb.save(str(out_path))
    return len(defects)
