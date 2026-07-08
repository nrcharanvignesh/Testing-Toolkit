"""
automation/report_excel.py
Generates Excel E2E test report using openpyxl.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .e2e_runner import TestCaseResult


# -- Style constants --
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_PASS_FONT = Font(color="006100")
_FAIL_FONT = Font(color="9C0006")


def _apply_header_style(ws, col_count: int) -> None:
    """Style header row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, col_count: int, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for col in range(1, col_count + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max(adjusted, 10)


def _add_conditional_formatting(ws, status_col: str, last_row: int) -> None:
    """Add green/red conditional formatting to status column."""
    cell_range = f"{status_col}2:{status_col}{last_row}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"pass"'],
            font=_PASS_FONT,
            fill=_PASS_FILL,
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"fail"'],
            font=_FAIL_FONT,
            fill=_FAIL_FILL,
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"error"'],
            font=_FAIL_FONT,
            fill=_FAIL_FILL,
        ),
    )


def write_e2e_report(
    results: list["TestCaseResult"],
    output_path: Path,
) -> Path:
    """Generate an Excel workbook with E2E test results.

    Sheets:
        Results: Per-step detail (TC ID, Step #, Action, Expected, Actual,
                 Status, Screenshot Path, Duration ms).
        Summary: Per-TC rollup (TC ID, Title, Overall Status, Total Steps,
                 Passed, Failed, Duration ms).

    Args:
        results: List of TestCaseResult from the runner.
        output_path: Path to write the .xlsx file.

    Returns:
        Path to the written workbook.
    """
    wb = Workbook()

    # -- Results sheet --
    ws_results = wb.active
    ws_results.title = "Results"  # type: ignore[union-attr]

    results_headers = [
        "TC ID", "Step #", "Action", "Expected", "Actual",
        "Status", "Screenshot Path", "Duration (ms)",
    ]
    ws_results.append(results_headers)

    for tc in results:
        for step in tc.steps:
            ws_results.append([
                tc.tc_id,
                step.step_num,
                step.action,
                step.expected,
                step.actual,
                step.status,
                str(step.screenshot_path) if step.screenshot_path else "",
                step.duration_ms,
            ])

    results_row_count = ws_results.max_row
    _apply_header_style(ws_results, len(results_headers))
    _auto_width(ws_results, len(results_headers))
    if results_row_count > 1:
        _add_conditional_formatting(ws_results, "F", results_row_count)

    # -- Summary sheet --
    ws_summary = wb.create_sheet("Summary")

    summary_headers = [
        "TC ID", "Title", "Overall Status", "Total Steps",
        "Passed", "Failed", "Duration (ms)",
    ]
    ws_summary.append(summary_headers)

    for tc in results:
        passed = sum(1 for s in tc.steps if s.status == "pass")
        failed = sum(1 for s in tc.steps if s.status in ("fail", "error"))
        ws_summary.append([
            tc.tc_id,
            tc.title,
            tc.overall_status,
            len(tc.steps),
            passed,
            failed,
            tc.duration_ms,
        ])

    summary_row_count = ws_summary.max_row
    _apply_header_style(ws_summary, len(summary_headers))
    _auto_width(ws_summary, len(summary_headers))
    if summary_row_count > 1:
        _add_conditional_formatting(ws_summary, "C", summary_row_count)

    # -- Save --
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
